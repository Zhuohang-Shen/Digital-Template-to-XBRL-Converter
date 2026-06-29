from __future__ import annotations

import logging
from collections import defaultdict
from datetime import date
from itertools import combinations
from typing import (
    TYPE_CHECKING,
    Optional,
)

if TYPE_CHECKING:
    from mireport.taxonomy import Concept, Taxonomy

from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.cell_range import CellRange

from mireport.conversionresults import ConversionResultsBuilder, MessageType, Severity
from mireport.xlsx_template_reader._bindings import (
    CellRangeMetadata,
    WorkbookBindings,
    XbrlConceptCellRangeMetadata,
    XbrlTableCellRangeMetadataHolder,
)
from mireport.xlsx_template_reader._cell_iteration import (
    getEffectiveCellRangeDimensions,
    getIteratorForCellRangeMetadata,
)
from mireport.xlsx_template_reader._constants import (
    EXCEL_PLACEHOLDER_VALUE,
    EXCEL_VALUES_TO_BE_TREATED_AS_NONE_VALUE,
    EXTERNAL_VALUES_RANGE,
    IGNORED_DEFINED_NAME_PREFIXES,
    CellType,
    CellValueType,
)
from mireport.xlsx_template_reader.util import (
    conceptsToText,
    excelCellRangeRef,
    excelCellRef,
    excelDefinedNameRef,
    getDateFromValue,
)

L = logging.getLogger(__name__)


class WorkbookReader:
    """Ergonomic cell-level access to an openpyxl Workbook.

    Carries the workbook and results builder, with internal tracking of
    unused named ranges.
    """

    def __init__(
        self,
        workbook: Workbook,
        results: ConversionResultsBuilder,
    ) -> None:
        self._workbook = workbook
        self._unused: set[DefinedName] = {
            dn
            for dn in workbook.defined_names.values()
            if (name := dn.name) and not name.startswith(IGNORED_DEFINED_NAME_PREFIXES)
        }
        self._results = results

    def close(self) -> None:
        self._workbook.close()

    def getDefinedName(self, name: str) -> Optional[DefinedName]:
        return self._workbook.defined_names.get(name)

    @property
    def unused_defined_names(self) -> frozenset[DefinedName]:
        return frozenset(self._unused)

    def build_bindings(self, taxonomy: Taxonomy, defaults: dict) -> WorkbookBindings:
        """Scrape named ranges from the workbook and return a WorkbookBindings."""
        concept_map: dict = {}
        unit_map: dict = {}
        preset_dims: defaultdict = defaultdict(dict)

        results = self._results

        for dn in self.unused_defined_names:
            concept = taxonomy.getConceptForName(dn.name)

            # TODO FIXME Temporary fix for the VSME taxonomy
            if dn.name == "IdentifierOfSitesInBiodiversitySensitiveAreasTypedAxis":
                concept = taxonomy.getConceptForName("IdentifierOfSiteTypedAxis")
            # TODO FIXME Temporary fix for the VSME taxonomy

            if concept is not None:
                if (crh := self._createCellRangeMetadata(dn)) is not None:
                    concept_map[dn] = (
                        XbrlConceptCellRangeMetadata.fromCellRangeMetadata(
                            crh, concept=concept
                        )
                    )
            elif "_" in dn.name:
                conceptName, _, memberName = dn.name.partition("_")
                if "unit" == memberName:
                    if (
                        concept := taxonomy.getConceptForName(conceptName)
                    ) is not None and (
                        crh := self._createCellRangeMetadata(dn)
                    ) is not None:
                        unit_map[concept] = (
                            XbrlConceptCellRangeMetadata.fromCellRangeMetadata(
                                crh, concept
                            )
                        )
                        self._unused.discard(dn)
                else:
                    concept = taxonomy.getConceptForName(conceptName)
                    dimValue = taxonomy.getConceptForName(memberName)
                    crh = self._createCellRangeMetadata(dn)
                    if crh is not None and concept is not None and dimValue is not None:
                        b = XbrlConceptCellRangeMetadata.fromCellRangeMetadata(
                            crh, concept=concept
                        )
                        if (
                            dim := taxonomy.getExplicitDimensionForDomainMember(
                                concept, dimValue
                            )
                        ) is not None:
                            concept_map[dn] = b
                            preset_dims[b][dim] = dimValue
                        else:
                            results.addMessage(
                                f"Domain member qualification set in named range {dn.name} but no dimension can be found for member.",
                                Severity.ERROR,
                                MessageType.DevInfo,
                            )
            if dn in concept_map:
                self._unused.discard(dn)

        results.addMessage(
            f"Excel file parsed ({results.numCellsPopulated} cells had data, with {results.numCellQueries} cells accessed).",
            Severity.INFO,
            MessageType.ExcelParsing,
        )

        table_map: dict = {}

        tables = [
            (dn, stuff)
            for dn, stuff in concept_map.items()
            if stuff.concept in taxonomy.hypercubes
        ]
        concepts_in_excel = frozenset(stuff.concept for stuff in concept_map.values())
        hc_concepts_in_excel = frozenset(c for c in concepts_in_excel if c.isHypercube)
        used_empty_hypercubes = taxonomy.emptyHypercubes.intersection(
            hc_concepts_in_excel
        )
        if used_empty_hypercubes:
            results.addMessage(
                f"The following hypercubes exist and have corresponding named ranges but they cannot be used due to missing taxonomy definitions: {conceptsToText(used_empty_hypercubes)}.",
                Severity.ERROR,
                MessageType.DevInfo,
            )

        for table, table_stuff in tables:
            tableCr = table_stuff.cellRange
            tableWorksheet = table_stuff.worksheet
            table_concept = table_stuff.concept

            allPermittedConceptsForTable = taxonomy.getDimensionsForHypercube(
                table_concept
            ).union(
                {
                    concept
                    for concept in taxonomy.getPrimaryItemsForHypercube(table_concept)
                    if concept.isReportable or concept.isDimension
                }
            )
            missing_from_excel = allPermittedConceptsForTable.difference(
                concepts_in_excel
            )
            if missing_from_excel:
                results.addMessage(
                    f"Expected Dimensions or Primary Items for hypercube {table.name} have not been found: {conceptsToText(missing_from_excel)}.",
                    Severity.WARNING,
                    MessageType.DevInfo,
                )

            candidates: list[XbrlConceptCellRangeMetadata] = []
            extras_in_excel: set[XbrlConceptCellRangeMetadata] = set()
            for dn, stuff in concept_map.items():
                if tableWorksheet is not stuff.worksheet:
                    continue
                concept = stuff.concept
                if not (concept.isReportable or concept.isDimension):
                    continue
                if tableCr.issuperset(stuff.cellRange):
                    if concept in allPermittedConceptsForTable:
                        candidates.append(stuff)
                    else:
                        extras_in_excel.add(stuff)
                elif not tableCr.isdisjoint(stuff.cellRange):
                    extras_in_excel.add(stuff)

            if extras_in_excel:
                results.addMessage(
                    f"Extra named ranges found within/overlapping bounds of {table.name} named range but not supported by Hypercube {table_stuff.concept.qname}: {extras_in_excel}.",
                    Severity.WARNING,
                    MessageType.DevInfo,
                )

            fishy = False
            for c1, c2 in combinations(candidates, 2):
                disjoint = c1.cellRange.isdisjoint(c2.cellRange)
                same = (
                    c1.concept.isReportable
                    and c2.concept.isReportable
                    and (c1.cellRange.bounds == c2.cellRange.bounds)
                )
                if not (disjoint or same):
                    fishy = True
                    results.addMessage(
                        f"Named range (table) {table.name} has named ranges (primary items or dimensions) {c1.definedName.name} and {c2.definedName.name} that are neither the same nor disjoint. Ignoring table.",
                        Severity.ERROR,
                        MessageType.ExcelParsing,
                    )
                    break

            if not fishy:
                pItems = [c for c in candidates if c.concept.isReportable]
                eDims = [c for c in candidates if c.concept.isExplicitDimension]
                tDims = [c for c in candidates if c.concept.isTypedDimension]
                units = [
                    u for p in pItems if (u := unit_map.get(p.concept)) is not None
                ]
                table_map[table_stuff] = XbrlTableCellRangeMetadataHolder(
                    primaryItems=pItems,
                    explicitDimensions=eDims,
                    typedDimensions=tDims,
                    units=units,
                )

        # Remove table entries from concept_map (they're now in table_map)
        for tableStuff, table_contents in table_map.items():
            concept_map.pop(tableStuff.definedName, None)
            table_dict = table_contents._asdict()
            for name, part_list in table_dict.items():
                for holder in part_list:
                    if "units" == name:
                        unit_map.pop(holder.concept, None)
                    else:
                        concept_map.pop(holder.definedName, None)

        has_external_value: set[Concept] = set()
        if (ext_dn := self._workbook.defined_names.get(EXTERNAL_VALUES_RANGE)) and (
            crh := self._createCellRangeMetadata(ext_dn)
        ):
            for cell in getIteratorForCellRangeMetadata(crh, only_cells=True):
                if not isinstance(cell.value, str):
                    continue
                name_or_label = cell.value.strip()
                if (
                    not name_or_label
                    or name_or_label in EXCEL_VALUES_TO_BE_TREATED_AS_NONE_VALUE
                ):
                    continue
                concept = taxonomy.getConceptForName(
                    name_or_label
                ) or taxonomy.getConceptForLabel(name_or_label)
                if concept is None or not concept.isTextblock:
                    self._results.addMessage(
                        f"External value specified in {EXTERNAL_VALUES_RANGE} named range but no matching concept found for name or label '{name_or_label}'.",
                        Severity.WARNING,
                        MessageType.DevInfo,
                        excel_reference=excelCellRef(crh.worksheet, cell),
                    )
                    continue
                has_external_value.add(concept)

        return WorkbookBindings(
            concept_map=concept_map,
            table_map=table_map,
            unit_map=unit_map,
            preset_dims=preset_dims,
            has_external_value=frozenset(has_external_value),
        )

    def _createCellRangeMetadata(self, dn: DefinedName) -> Optional[CellRangeMetadata]:
        try:
            all_destinations = list(dn.destinations)
        except AttributeError:
            self._results.addMessage(
                f"Named range {dn.name} has an unreadable destination: {dn.attr_text!r}. \nSomething has modified the digital template's structure. \nPlease try a fresh copy of the template and check that it has not been modified in unsupported ways.",
                Severity.ERROR,
                MessageType.DevInfo,
            )
            L.exception(
                f"OpenPyXL error processing named range definition {dn.name=} {dn.attr_text=!r}."
            )
            return None
        match len(all_destinations):
            case 0:
                self._results.addMessage(
                    f"Named range {dn.name} has no destinations specified. Ignoring.",
                    Severity.ERROR,
                    MessageType.DevInfo,
                )
                return None
            case 1:
                pass
            case _:
                self._results.addMessage(
                    f"Table {dn.name} has multiple destinations. Ignoring table.",
                    Severity.ERROR,
                    MessageType.DevInfo,
                )
                return None
        sheetName, cell_range = all_destinations[0]
        if not sheetName or not cell_range:
            self._results.addMessage(
                f"Named range {dn.name} has damaged cell reference {sheetName=} {cell_range=}",
                Severity.ERROR,
                MessageType.ExcelParsing,
            )
            return None
        try:
            ws = self._workbook[sheetName]
            cr = CellRange(cell_range)
        except Exception as e:
            L.exception(
                f"OpenPyXL error processing cell range. {dn.name=} {sheetName=} {cell_range=}",
                exc_info=e,
            )
            return None
        dims = getEffectiveCellRangeDimensions(ws, cr)
        self._results.addCellQueries(dims.cellsAccessed)
        self._results.addCellsWithData(dims.cellsPopulated)
        return CellRangeMetadata(
            dn,
            ws,
            cr,
            populated_height=dims.populated_height,
            populated_width=dims.populated_width,
            populated_min_col=dims.populated_min_col,
            populated_min_row=dims.populated_min_row,
        )

    def _getCellRangeMetadata(
        self,
        definedName: DefinedName
        | str
        | XbrlConceptCellRangeMetadata
        | CellRangeMetadata,
    ) -> Optional[CellRangeMetadata]:
        if isinstance(definedName, str):
            definedName = self._workbook.defined_names.get(definedName)
            if definedName is None:
                return None
        if isinstance(definedName, DefinedName):
            if (crm := self._createCellRangeMetadata(definedName)) is None:
                return None
            definedName = crm
        if isinstance(definedName, (XbrlConceptCellRangeMetadata, CellRangeMetadata)):
            self._unused.discard(definedName.definedName)
            return definedName
        return None

    def getSingleCell(
        self,
        definedName: DefinedName
        | str
        | XbrlConceptCellRangeMetadata
        | CellRangeMetadata,
        *,
        row: int = -1,
        column: int = -1,
    ) -> Optional[CellType]:
        if (stuff := self._getCellRangeMetadata(definedName)) is None:
            return None

        cr = stuff.cellRange
        ws = stuff.worksheet

        if not all(
            x is not None for x in (cr.min_row, cr.max_row, cr.min_col, cr.max_col)
        ):
            self._results.addMessage(
                f"Named range {stuff.definedName.name} has an invalid cell range {cr.bounds}.",
                Severity.ERROR,
                MessageType.DevInfo,
                excel_reference=excelDefinedNameRef(stuff.definedName),
            )
            return None

        shouldOverrideRow = row == -1 or stuff.maximum_height == 1
        shouldOverrideColumn = column == -1 or stuff.maximum_width == 1

        if shouldOverrideRow:
            row = cr.min_row
            if stuff.populated_height > 1:
                self._results.addMessage(
                    f"Named range {stuff.definedName.name} has {stuff.populated_height} populated rows but no row was specified; using the first.",
                    Severity.WARNING,
                    MessageType.DevInfo,
                    excel_reference=excelCellRangeRef(ws, cr),
                )

        if shouldOverrideColumn:
            column = cr.min_col
            if stuff.populated_width > 1:
                self._results.addMessage(
                    f"Named range {stuff.definedName.name} has {stuff.populated_width} populated columns but no column was specified; using the first.",
                    Severity.WARNING,
                    MessageType.DevInfo,
                    excel_reference=excelCellRangeRef(ws, cr),
                )

        if not (cr.min_row <= row <= cr.max_row):
            self._results.addMessage(
                f"Row {row} has not been specified correctly.",
                Severity.WARNING,
                MessageType.DevInfo,
                excel_reference=excelCellRangeRef(ws, cr),
            )
            row = cr.min_row
        if not (cr.min_col <= column <= cr.max_col):
            self._results.addMessage(
                f"Column {column} has not been specified correctly.",
                Severity.WARNING,
                MessageType.DevInfo,
                excel_reference=excelCellRangeRef(ws, cr),
            )
            column = cr.min_col

        cell = ws.cell(row=row, column=column)

        if cell is None or cell.value is None:
            return None

        if cell.value == EXCEL_PLACEHOLDER_VALUE:
            self._results.addMessage(
                f"Excel cell has an invalid stored value {EXCEL_PLACEHOLDER_VALUE}. Please check the Excel formula for this specific cell.",
                Severity.ERROR,
                MessageType.ExcelParsing,
                excel_reference=excelCellRef(ws, cell),
            )
            return None
        return cell

    def getSingleValue(
        self,
        definedName: DefinedName | str,
        *,
        row: int = -1,
        column: int = -1,
    ) -> CellValueType:
        if (
            cell := self.getSingleCell(definedName, row=row, column=column)
        ) is not None:
            value = cell.value
            if not isinstance(value, CellValueType):
                value = str(value)
            return value
        return None

    def getSingleStringValue(
        self,
        definedName: DefinedName | str,
        *,
        row: int = -1,
        column: int = -1,
        fallbackValue: str = "",
    ) -> str:
        value = self.getSingleValue(definedName, row=row, column=column)
        return str(value) if value is not None else str(fallbackValue)

    def getSingleDateValue(self, definedName: DefinedName | str) -> date:
        value = self.getSingleValue(definedName)
        return getDateFromValue(value)
