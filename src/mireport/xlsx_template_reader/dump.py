from __future__ import annotations

from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.cell_range import CellRange

from mireport.exceptions import OpenPyXlRelatedException
from mireport.xlsx_template_reader._constants import CellValueType


class NamedRangeException(OpenPyXlRelatedException):
    """Exception raised when a named range is broken in the workbook."""

    def __init__(self, message: str, defined_name: DefinedName) -> None:
        self.message = message
        self.defined_name = defined_name
        super().__init__(message, str(defined_name))

    def __str__(self) -> str:
        details = (
            f"Details:\n"
            f"  Name: {self.defined_name.name}\n"
            f"  Refers to: {self.defined_name.attr_text}\n"
        )
        return f"{self.message} {details}"


def list_named_ranges(wb: Workbook) -> list[tuple[str, str]]:
    """Return (name, raw_value) for every defined name in the workbook, in iteration order."""
    return [(dn.name, dn.attr_text) for dn in wb.defined_names.values()]


def getNamedRanges(
    wb: Workbook,
) -> tuple[dict[str, list[CellValueType]], list[NamedRangeException]]:
    data = {}
    errors = []
    for dn in list(wb.defined_names.values()):
        if not dn.name:
            errors.append(
                NamedRangeException("Named range exists but has no name.", dn)
            )
            continue

        if not dn.destinations:
            errors.append(
                NamedRangeException("Named range has no destination specified.", dn)
            )
            continue

        try:
            parts = next(iter(dn.destinations))
        except (ValueError, AttributeError):
            errors.append(
                NamedRangeException("Named range has invalid destinations.", dn)
            )
            continue

        sheet_name, cell_range = parts
        if sheet_name not in wb:
            errors.append(
                NamedRangeException(
                    "Named range refers to a worksheet that is not in the workbook.", dn
                )
            )
            continue

        ws = wb[sheet_name]

        if not cell_range:
            errors.append(
                NamedRangeException("Named range has no cell range specified.", dn)
            )
            continue

        cr = CellRange(cell_range)

        if (
            cr.min_col is None
            or cr.min_row is None
            or cr.max_col is None
            or cr.max_row is None
        ):
            errors.append(
                NamedRangeException(
                    f"Named range cell range bounds expected to be int but actually None {cr=}.",
                    dn,
                )
            )
            continue

        width: int = cr.max_col - cr.min_col
        height: int = cr.max_row - cr.min_row
        if width < 0 or height < 0:
            errors.append(
                NamedRangeException(
                    f"Named range has negative cell range {width=} {height=}.", dn
                )
            )
            continue

        if not width and not height:
            # a single (width=0, height=0) cell range ... so the OpenPyXL API returns it directly.
            cell = ws[cell_range]
            values = [cell.value]
        else:
            values = []
            for row in ws[cell_range]:
                values.extend([c.value for c in row])
        data[dn.name] = values

    return data, errors
